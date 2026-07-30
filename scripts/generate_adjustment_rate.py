#!/usr/bin/env python3
"""生成调价率模块数据。

整体调价率：销售看板 data 底表，严格复刻“总览”TOTAL 行口径：
B=小组、BT=cur、I=全部(*)，按 H 小组汇总 BK/BL。
六高调价率：六高“用户关注商品明细”，以调价状态=已调价/未调价为有效分母，
分母剔除与 NULL 不进入统计。
"""
from __future__ import annotations
import argparse
import datetime as dt
import json
from pathlib import Path
import openpyxl

GROUPS = ['饰品1组','饰品2组','海淘组','珠宝1组','珠宝2组','珠宝3组']
OVERALL_TARGET = 0.75
SIX_HIGH_TARGET = 0.80


def as_date(v):
    s = str(v or '').split('.')[0]
    if len(s) == 8 and s.isdigit():
        return f'{s[:4]}-{s[4:6]}-{s[6:]}'
    return None


def overall(source: Path):
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    if 'data' not in wb.sheetnames or 'brand_info' not in wb.sheetnames:
        raise RuntimeError('销售看板缺少 data/brand_info 子表')
    data = wb['data']
    source_date = as_date(wb['brand_info']['D2'].value)
    result = {g: {'denominator': 0.0, 'adjusted': 0.0} for g in GROUPS}
    for row in data.iter_rows(min_row=2, values_only=True):
        # B 数据分组、H 小组、I 神银、BK 价高商品数、BL 调价数、BT cur 标识
        if row[1] != '小组' or row[7] not in result or row[71] != 'cur' or row[8] in (None, ''):
            continue
        result[row[7]]['denominator'] += float(row[62] or 0)
        result[row[7]]['adjusted'] += float(row[63] or 0)
    wb.close()
    return source_date, result


def six_high(source: Path):
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    if '用户关注商品明细' not in wb.sheetnames:
        raise RuntimeError('六高报表缺少“用户关注商品明细”子表')
    ws = wb['用户关注商品明细']
    headers = [c.value for c in ws[2]]
    required = ['日期', '小组', '调价状态']
    missing = [x for x in required if x not in headers]
    if missing:
        raise RuntimeError(f'六高明细缺少字段: {missing}')
    idx = {name: headers.index(name) for name in required}
    result = {g: {'denominator': 0, 'adjusted': 0} for g in GROUPS}
    dates = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        group, status = row[idx['小组']], row[idx['调价状态']]
        d = as_date(row[idx['日期']])
        if d: dates.append(d)
        if group not in result or status not in ('已调价', '未调价'):
            continue
        result[group]['denominator'] += 1
        if status == '已调价':
            result[group]['adjusted'] += 1
    wb.close()
    return max(dates) if dates else None, result


def normalize(group, overall_counts, six_counts):
    def block(counts, target):
        den = int(round(counts['denominator']))
        adj = int(round(counts['adjusted']))
        rate = adj / den if den else None
        return {'denominator': den, 'adjusted': adj, 'rate': rate, 'target': target,
                'status': 'no_data' if rate is None else ('pass' if rate >= target else 'fail')}
    return {'group': group, 'overall': block(overall_counts, OVERALL_TARGET),
            'six_high': block(six_counts, SIX_HIGH_TARGET)}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sales', required=True, type=Path)
    ap.add_argument('--six-high', required=True, type=Path)
    ap.add_argument('--output', required=True, type=Path)
    args = ap.parse_args()
    overall_date, ov = overall(args.sales)
    six_date, sh = six_high(args.six_high)
    rows = [normalize(g, ov[g], sh[g]) for g in GROUPS]
    ov_total = {k: sum(ov[g][k] for g in GROUPS) for k in ('denominator','adjusted')}
    sh_total = {k: sum(sh[g][k] for g in GROUPS) for k in ('denominator','adjusted')}
    summary = normalize('精品总', ov_total, sh_total)

    errors = []
    for row in rows + [summary]:
        for key in ('overall','six_high'):
            b = row[key]
            if b['adjusted'] > b['denominator']:
                errors.append(f"{row['group']} {key} 分子大于分母")
            if b['rate'] is not None and abs(b['rate'] - b['adjusted']/b['denominator']) > 1e-12:
                errors.append(f"{row['group']} {key} 比率不一致")
    if summary['overall']['denominator'] != sum(x['overall']['denominator'] for x in rows):
        errors.append('整体精品总分母不等于小组合计')
    if summary['six_high']['denominator'] != sum(x['six_high']['denominator'] for x in rows):
        errors.append('六高精品总分母不等于小组合计')
    if errors:
        raise RuntimeError('; '.join(errors))

    payload = {
        'title': '调价率',
        'overall_source_date': overall_date,
        'six_high_source_date': six_date,
        'overall_target': OVERALL_TARGET,
        'six_high_target': SIX_HIGH_TARGET,
        'groups': rows,
        'summary': summary,
        'generated_at': dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'validation': {'status': 'PASS', 'errors': []},
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    print(json.dumps(payload, ensure_ascii=False, indent=2))

if __name__ == '__main__':
    main()
