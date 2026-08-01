#!/usr/bin/env python3
"""将销售看板清洗为品牌日明细标准层，再生成前端品牌调价率月历。

关键修正：指标大表使用 xlsx XML 流式读取，而不是 openpyxl read_only。
本工作簿 `data非神银` 后段存在 openpyxl read_only 漏读风险，礼兰 SN 10202821
就在后段；必须直接读取原始 worksheet XML 才能保证完整。
"""
from __future__ import annotations
import argparse, calendar, datetime as dt, hashlib, json, re, zipfile, xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path
import openpyxl

SOURCE_SHEETS = ('data', 'data非神银')
NULLS = ('', 'NULL', '(NULL)')
NS = {'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main', 'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}


def text(v):
    s = str(v or '').strip()
    return s[:-2] if s.endswith('.0') else s


def date_text(v):
    s = text(v)
    return f'{s[:4]}-{s[4:6]}-{s[6:]}' if re.fullmatch(r'\d{8}', s) else s


def col(headers, name):
    for i, value in enumerate(headers):
        if str(value or '').strip() == name:
            return i
    raise RuntimeError(f'缺少字段 {name}')


def number(v):
    if v in (None, ''):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError) as exc:
        raise RuntimeError(f'指标不是数值: {v!r}') from exc


def load_catalog(wb):
    ws = wb['brand_info']
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {x: col(headers, x) for x in ('日期', '品牌', 'sn', '等级', '小组', '神银')}
    catalog, conflicts = {}, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sn = text(row[idx['sn']])
        if sn in NULLS:
            continue
        item = {
            'sn': sn,
            'brand': str(row[idx['品牌']] or '').strip(),
            'level': str(row[idx['等级']] or '').strip(),
            'group': str(row[idx['小组']] or '').strip(),
            'groups': [],
            'shenyin': str(row[idx['神银']] or '').strip(),
            'info_date': date_text(row[idx['日期']]),
        }
        if sn not in catalog:
            catalog[sn] = item
        else:
            old = catalog[sn]
            if any(old[k] != item[k] for k in ('brand', 'level', 'shenyin')):
                conflicts.append({'sn': sn, 'old': old, 'new': item})
        group = item['group']
        if group and group not in catalog[sn]['groups']:
            catalog[sn]['groups'].append(group)
    if conflicts:
        raise RuntimeError(f'品牌主数据冲突: {conflicts[:5]}')
    return catalog


def shared_strings(zf):
    if 'xl/sharedStrings.xml' not in zf.namelist():
        return []
    root = ET.fromstring(zf.read('xl/sharedStrings.xml'))
    out = []
    for si in root.findall('a:si', NS):
        parts = [t.text or '' for t in si.findall('.//a:t', NS)]
        out.append(''.join(parts))
    return out


def workbook_sheet_paths(zf):
    wb_root = ET.fromstring(zf.read('xl/workbook.xml'))
    rel_root = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
    rels = {rel.attrib['Id']: rel.attrib['Target'] for rel in rel_root}
    paths = {}
    for sh in wb_root.findall('.//a:sheet', NS):
        name = sh.attrib['name']
        rid = sh.attrib['{%s}id' % NS['r']]
        target = rels[rid]
        paths[name] = 'xl/' + target.lstrip('/') if not target.startswith('xl/') else target
    return paths


def column_index(cell_ref):
    m = re.match(r'([A-Z]+)', cell_ref)
    letters = m.group(1)
    n = 0
    for ch in letters:
        n = n * 26 + ord(ch) - 64
    return n - 1


def cell_value(cell, sst):
    t = cell.attrib.get('t')
    if t == 'inlineStr':
        return ''.join(x.text or '' for x in cell.findall('.//a:t', NS))
    v = cell.find('a:v', NS)
    if v is None:
        return ''
    raw = v.text or ''
    if t == 's':
        return sst[int(raw)]
    return raw


def iter_sheet_rows_from_xml(source: Path, sheet_name: str):
    with zipfile.ZipFile(source) as zf:
        sst = shared_strings(zf)
        paths = workbook_sheet_paths(zf)
        if sheet_name not in paths:
            raise RuntimeError(f'缺少子表: {sheet_name}')
        with zf.open(paths[sheet_name]) as fh:
            for event, elem in ET.iterparse(fh, events=('end',)):
                if elem.tag.endswith('}row'):
                    values = []
                    for c in elem.findall('a:c', NS):
                        idx = column_index(c.attrib['r'])
                        if idx >= len(values):
                            values.extend([''] * (idx + 1 - len(values)))
                        values[idx] = cell_value(c, sst)
                    yield values
                    elem.clear()


def clean_detail(source: Path, catalog):
    aggregate = defaultdict(lambda: {
        'high_price_count': 0.0,
        'adjusted_count': 0.0,
        'source_row_count': 0,
        'source_sheets': set(),
    })
    sheet_brand_sets, valid_dates = {}, []
    missing_catalog = {}
    for sheet_name in SOURCE_SHEETS:
        rows = iter_sheet_rows_from_xml(source, sheet_name)
        headers = next(rows)
        idx = {x: col(headers, x) for x in (
            '数据分组\n（选TOTAL来用）', '日期', 'sn', '品牌', '等级', '小组', '神银',
            '价高商品数', '价高调价商品数', '辅助列3')}
        brands = set()
        for row in rows:
            if len(row) <= max(idx.values()):
                continue
            if row[idx['数据分组\n（选TOTAL来用）']] != 'TOTAL' or row[idx['辅助列3']] != 'cur':
                continue
            sn, raw_date = text(row[idx['sn']]), text(row[idx['日期']])
            if sn in NULLS or not re.fullmatch(r'\d{8}', raw_date):
                continue
            den = number(row[idx['价高商品数']])
            adj = number(row[idx['价高调价商品数']])
            if den < 0 or adj < 0 or adj > den:
                raise RuntimeError(f'{sheet_name} {sn} {raw_date} 指标异常: {adj}>{den}')
            key = (sn, raw_date)
            aggregate[key]['high_price_count'] += den
            aggregate[key]['adjusted_count'] += adj
            aggregate[key]['source_row_count'] += 1
            aggregate[key]['source_sheets'].add(sheet_name)
            brands.add(sn)
            valid_dates.append(raw_date)
            if sn not in catalog:
                group = str(row[idx['小组']] or '').strip()
                missing_catalog[sn] = {
                    'sn': sn, 'brand': str(row[idx['品牌']] or '').strip(),
                    'level': str(row[idx['等级']] or '').strip(),
                    'group': group,
                    'groups': [group] if group else [],
                    'shenyin': str(row[idx['神银']] or '').strip(),
                    'info_date': date_text(raw_date),
                }
        sheet_brand_sets[sheet_name] = brands
    if not valid_dates:
        raise RuntimeError('两张指标表均未找到 TOTAL + cur 有效记录')
    overlap = sorted(sheet_brand_sets['data'] & sheet_brand_sets['data非神银'])
    if overlap:
        raise RuntimeError(f'指标工作表品牌SN重叠，禁止重复累计: {overlap[:20]}')
    catalog.update(missing_catalog)
    latest = max(valid_dates)
    month = latest[:6]
    records = []
    for (sn, raw_date), value in sorted(aggregate.items(), key=lambda x: (x[0][1], x[0][0])):
        if not raw_date.startswith(month):
            continue
        den = int(round(value['high_price_count']))
        adj = int(round(value['adjusted_count']))
        meta = catalog[sn]
        records.append({
            'date': date_text(raw_date), 'brand_sn': sn, 'brand_name': meta['brand'],
            'group': meta['group'], 'level': meta['level'], 'shenyin_type': meta['shenyin'],
            'high_price_count': den, 'adjusted_count': adj,
            'rate': adj / den if den else None,
            'source_sheets': sorted(value['source_sheets']),
            'source_row_count': value['source_row_count'],
        })
    return records, catalog, latest, month, sheet_brand_sets


def build_front(records, catalog, latest, month):
    by_brand = defaultdict(lambda: {'denominator': 0, 'adjusted': 0, 'row_count': 0})
    by_day = defaultdict(dict)
    for row in records:
        sn = row['brand_sn']
        by_brand[sn]['denominator'] += row['high_price_count']
        by_brand[sn]['adjusted'] += row['adjusted_count']
        by_brand[sn]['row_count'] += row['source_row_count']
        by_day[sn][row['date'].replace('-', '')] = row
    year, mon = int(month[:4]), int(month[4:6])
    month_days = calendar.monthrange(year, mon)[1]
    brands = []
    for sn, meta in catalog.items():
        total = by_brand[sn]
        daily = []
        for day in range(1, month_days + 1):
            raw_date = f'{month}{day:02d}'
            row = by_day[sn].get(raw_date)
            daily.append({
                'date': date_text(raw_date), 'day': day,
                'denominator': row['high_price_count'] if row else 0,
                'adjusted': row['adjusted_count'] if row else 0,
                'rate': row['rate'] if row else None,
                'has_data': row is not None,
                'is_future': raw_date > latest,
            })
        den, adj = total['denominator'], total['adjusted']
        brands.append({
            **meta, 'date': date_text(latest), 'month': f'{mon}月截至{int(latest[6:])}日',
            'denominator': den, 'adjusted': adj, 'rate': adj / den if den else None,
            'row_count': total['row_count'], 'daily': daily,
        })
    brands.sort(key=lambda x: (-x['denominator'], x['brand'], x['sn']))
    return brands


def validate(records, brands, catalog, sheet_brand_sets, latest, month):
    errors = []
    record_keys = [(x['brand_sn'], x['date']) for x in records]
    if len(record_keys) != len(set(record_keys)):
        errors.append('标准明细品牌SN+日期不唯一')
    brand_sns = [x['sn'] for x in brands]
    if len(brand_sns) != len(set(brand_sns)):
        errors.append('前端品牌SN不唯一')
    for row in records:
        if row['adjusted_count'] > row['high_price_count']:
            errors.append(f"{row['brand_sn']} {row['date']} 分子大于分母")
        if not row['date'].replace('-', '').startswith(month):
            errors.append(f"{row['brand_sn']} 日期越界")
    for brand in brands:
        if brand['denominator'] != sum(x['denominator'] for x in brand['daily']):
            errors.append(f"{brand['sn']} 月价高数与日明细不一致")
        if brand['adjusted'] != sum(x['adjusted'] for x in brand['daily']):
            errors.append(f"{brand['sn']} 月调价数与日明细不一致")
    overlap = sheet_brand_sets['data'] & sheet_brand_sets['data非神银']
    indicator_count = len(sheet_brand_sets['data'] | sheet_brand_sets['data非神银'])
    if overlap:
        errors.append('两张指标表品牌重叠')
    if indicator_count < 500:
        errors.append(f'指标品牌覆盖异常下降: {indicator_count}<500')
    regression = {x['sn']: x for x in brands}
    if regression.get('10202821', {}).get('row_count', 0) < 30:
        errors.append('回归失败：礼兰 SN 10202821 应有30天非神银明细')
    # 月度源数据会随日期推进变化；不再用旧日期固定数值断言（例如 CACUSS=119）阻断月末更新。
    # 保留上方的唯一性、分子<=分母、月汇总=日明细、双表不重叠和品牌覆盖率校验作为发布门禁。
    if len(catalog) < indicator_count:
        errors.append('品牌目录少于指标品牌数')
    if errors:
        raise RuntimeError('; '.join(errors[:20]))
    return {
        'status': 'PASS', 'errors': [], 'latest_date': date_text(latest),
        'catalog_count': len(catalog), 'indicator_brand_count': indicator_count,
        'metric_brand_count': sum(x['denominator'] > 0 for x in brands),
        'standard_record_count': len(records),
        'source_brand_counts': {k: len(v) for k, v in sheet_brand_sets.items()},
        'source_overlap_count': len(overlap),
        'reader': 'xlsx-xml-stream',
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--source', required=True, type=Path)
    ap.add_argument('--daily-output', required=True, type=Path)
    ap.add_argument('--output', required=True, type=Path)
    args = ap.parse_args()
    wb = openpyxl.load_workbook(args.source, data_only=True, read_only=False)
    for sheet in ('brand_info', *SOURCE_SHEETS):
        if sheet not in wb.sheetnames:
            raise RuntimeError(f'缺少子表: {sheet}')
    catalog = load_catalog(wb)
    wb.close()
    records, catalog, latest, month, sheet_brand_sets = clean_detail(args.source, catalog)
    brands = build_front(records, catalog, latest, month)
    quality = validate(records, brands, catalog, sheet_brand_sets, latest, month)
    source_hash = hashlib.sha256(args.source.read_bytes()).hexdigest()
    generated_at = dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    daily_payload = {
        'title': '品牌调价率标准日明细', 'schema_version': '1.1',
        'source_file': args.source.name, 'source_sha256': source_hash,
        'source_date': date_text(latest), 'source_month': month,
        'grain': 'brand_sn + date',
        'filters': {'data_group': 'TOTAL', 'period_flag': 'cur', 'source_sheets': list(SOURCE_SHEETS)},
        'fields': ['date', 'brand_sn', 'brand_name', 'group', 'level', 'shenyin_type', 'high_price_count', 'adjusted_count', 'rate', 'source_sheets', 'source_row_count'],
        'records': records, 'quality': quality, 'generated_at': generated_at,
    }
    front_payload = {
        'title': '品牌调价率', 'source_date': date_text(latest), 'source_month': month,
        'scope': 'data + data非神银；数据分组=TOTAL；辅助列3=cur；标准层按品牌SN+日期汇总BK/BL；使用xlsx XML流式读取确保大表完整。',
        'coverage_note': '品牌目录来自brand_info；指标来自data与data非神银，先清洗为品牌日明细再生成月历。',
        'brands': brands,
        'summary': {
            'catalog_count': len(brands),
            'indicator_brand_count': quality['indicator_brand_count'],
            'metric_brand_count': quality['metric_brand_count'],
            'denominator': sum(x['denominator'] for x in brands),
            'adjusted': sum(x['adjusted'] for x in brands),
        },
        'validation': quality, 'generated_at': generated_at,
    }
    args.daily_output.parent.mkdir(parents=True, exist_ok=True)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.daily_output.write_text(json.dumps(daily_payload, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    args.output.write_text(json.dumps(front_payload, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    print(json.dumps({'daily': quality, 'front_summary': front_payload['summary']}, ensure_ascii=False))


if __name__ == '__main__':
    main()
