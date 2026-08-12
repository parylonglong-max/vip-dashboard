#!/usr/bin/env python3
"""品牌分层MTD数据生成脚本

数据源：
  - VMA 下载的销售+流量 Excel（品牌粒度，含小组分组）
  - 货价监控指标-26.8.10.xlsx / 品牌分层数据 O-T 列（品牌分层映射）

输出：
  - dashboard_project/frontend/data/brand_tier_mtd.json（前端数据）
  - 品牌分层MTD_复核报告_YYYYMMDD.json（审计报告）

规则：
  - 品牌SN是唯一匹配主键
  - 历史小组（饰品3组、手表组等）直接排除
  - 未在映射表中找到但有数据的品牌：出现在品牌视角明细，不参与汇总表
  - 无数据品牌：完全不出现
  - 汇总由底层明细重算，占比=分子/总分母，同比=本期/同期-1
"""
from __future__ import annotations

import argparse
import json
import sys
import datetime as dt
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

# ── 常量 ─────────────────────────────────────────────────

ACTIVE_GROUPS = ['饰品1组', '饰品2组', '海淘组', '珠宝1组', '珠宝2组', '珠宝3组']
HISTORICAL_GROUPS = ['饰品3组', '手表2组（20240708取消）', '手表组（20250220取消）']

TIER_ORDER = ['S1', 'S2', 'S3', '高价值', '矩阵非高', '双非']
CATEGORY_MAP = {
    'S1': '标品', 'S2': '标品', 'S3': '标品',
    '高价值': '类穿戴', '矩阵非高': '类穿戴', '双非': '类穿戴',
}
CATEGORY_ORDER = ['标品', '类穿戴']

OUTPUT_FILE = 'dashboard_project/frontend/data/brand_tier_mtd.json'
AUDIT_FILE = '品牌分层MTD_复核报告_${date}.json'
DEFAULT_MAPPING_SOURCE = '货价监控指标-26.8.10.xlsx'


# ── 工具函数 ─────────────────────────────────────────────

def parse_num(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v if v == v else None  # NaN check
    if isinstance(v, str):
        s = v.strip()
        if not s or s.lower() in ('', '(null)', 'null', '#value!', '#n/a', '#div/0!', 'nan'):
            return None
        try:
            return float(s.replace('%', '').replace(',', ''))
        except ValueError:
            return None
    return None


def format_unit(v: Optional[float], decimals_wan: int = 1, decimals_yi: int = 2) -> str:
    """自动切换单位：亿/万/原值"""
    if v is None:
        return '—'
    abs_v = abs(v)
    if abs_v >= 100_000_000:
        return f'{v / 100_000_000:.{decimals_yi}f}亿'
    elif abs_v >= 10_000:
        return f'{v / 10_000:.{decimals_wan}f}万'
    else:
        return f'{int(v)}'


def safe_div(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if a is None or b is None or b == 0:
        return None
    return a / b


def calc_yoy(current: Optional[float], compare: Optional[float]) -> Optional[float]:
    """计算同比：本期/同期-1"""
    if current is None or compare is None or compare == 0:
        return None
    return current / compare - 1


# ── 读取数据 ─────────────────────────────────────────────

def read_vma_data(vma_path: str) -> pd.DataFrame:
    """读取 VMA 下载的销售+流量数据"""
    df = pd.read_excel(vma_path, dtype={'品牌SN': str})

    # 标准化列名
    df.columns = df.columns.str.strip()

    # 数值列
    num_cols = ['销售额(含拒退)', '销售额(含拒退)(对比)', '累计曝光流量', '累计曝光流量(对比)']
    for c in num_cols:
        df[c] = pd.to_numeric(df[c], errors='coerce')

    # 排除历史小组
    before = len(df)
    df = df[~df['二级业绩部类'].isin(HISTORICAL_GROUPS)]
    print(f'  排除历史小组: {before} → {len(df)} 行 ({", ".join(HISTORICAL_GROUPS)})')

    # 只保留有效小组
    df = df[df['二级业绩部类'].isin(ACTIVE_GROUPS)]
    print(f'  保留有效小组: {len(df)} 行')

    return df


def read_tier_mapping(mapping_source: str) -> Dict[str, Dict[str, str]]:
    """从货价监控表 O-T 列读取品牌分层映射"""
    wb = load_workbook(mapping_source, data_only=True)
    ws = wb['品牌分层数据']

    mapping: Dict[str, Dict[str, str]] = {}
    for i in range(2, 1000):
        sn = str(ws.cell(i, 15).value)
        tier = ws.cell(i, 20).value
        brand = ws.cell(i, 19).value
        cat = str(ws.cell(i, 17).value or '').strip()
        if sn is None or sn == 'None' or sn == '':
            break
        if tier is None or str(tier).strip() == '':
            continue
        tier = str(tier).strip()
        # 类别推断
        if cat in ('标品', '类穿戴'):
            category = cat
        else:
            category = CATEGORY_MAP.get(tier, '')
        mapping[sn] = {'tier': tier, 'category': category, 'brand': brand or ''}

    wb.close()
    print(f'  品牌分层映射: {len(mapping)} 个唯一SN')
    tier_dist = {}
    for v in mapping.values():
        tier_dist[v['tier']] = tier_dist.get(v['tier'], 0) + 1
    print(f'  分层分布: {dict(sorted(tier_dist.items()))}')
    return mapping


def calc_summary(
    df: pd.DataFrame,
    mapping: Dict[str, Dict[str, str]],
    matched_only: bool = True,
) -> List[Dict[str, Any]]:
    """计算品牌分层汇总表

    当 matched_only=True 时，仅统计有映射标签的品牌（用于汇总表）
    当 matched_only=False 时，统计所有品牌（用于审计对比）
    """
    result = []

    # 过滤
    if matched_only:
        work_df = df[df['品牌SN'].isin(mapping)].copy()
    else:
        work_df = df.copy()

    # 分层标签
    if matched_only:
        work_df['tier'] = work_df['品牌SN'].map(lambda sn: mapping.get(sn, {}).get('tier', ''))
    else:
        # 未匹配的标为 '(无标签)'
        work_df['tier'] = work_df['品牌SN'].map(
            lambda sn: mapping.get(sn, {}).get('tier', '(无标签)')
        )
    work_df['category'] = work_df['tier'].map(CATEGORY_MAP.get)

    total_sales = float(work_df['销售额(含拒退)'].sum() or 0)
    total_traffic = float(work_df['累计曝光流量'].sum() or 0)
    total_sales_compare = float(work_df['销售额(含拒退)(对比)'].sum() or 0)
    total_traffic_compare = float(work_df['累计曝光流量(对比)'].sum() or 0)

    # 按分层汇总
    tier_data = {}
    for tier in TIER_ORDER:
        tdf = work_df[work_df['tier'] == tier]
        if len(tdf) == 0:
            continue
        sales = float(tdf['销售额(含拒退)'].sum() or 0)
        sales_compare = float(tdf['销售额(含拒退)(对比)'].sum() or 0)
        traffic = float(tdf['累计曝光流量'].sum() or 0)
        traffic_compare = float(tdf['累计曝光流量(对比)'].sum() or 0)
        tier_data[tier] = {
            'sales': sales, 'sales_compare': sales_compare,
            'traffic': traffic, 'traffic_compare': traffic_compare,
        }

    # 分类汇总
    for cat in CATEGORY_ORDER:
        cat_tiers = [t for t in TIER_ORDER if CATEGORY_MAP.get(t) == cat]
        cat_sum = {'sales': 0, 'sales_compare': 0, 'traffic': 0, 'traffic_compare': 0}
        for t in cat_tiers:
            if t in tier_data:
                for k in cat_sum:
                    cat_sum[k] += tier_data[t][k]

        for t in cat_tiers:
            if t not in tier_data:
                continue
            d = tier_data[t]
            result.append({
                'category': cat,
                'tier': t,
                'sales': d['sales'],
                'sales_share': safe_div(d['sales'], total_sales) if total_sales else None,
                'sales_compare': d['sales_compare'],
                'sales_yoy': calc_yoy(d['sales'], d['sales_compare']),
                'traffic': d['traffic'],
                'traffic_share': safe_div(d['traffic'], total_traffic) if total_traffic else None,
                'traffic_compare': d['traffic_compare'],
                'traffic_yoy': calc_yoy(d['traffic'], d['traffic_compare']),
                'is_subtotal': False,
            })

        # 分类合计
        if cat_sum['sales'] > 0 or cat_sum['traffic'] > 0:
            result.append({
                'category': cat,
                'tier': '合计',
                'sales': cat_sum['sales'],
                'sales_share': safe_div(cat_sum['sales'], total_sales) if total_sales else None,
                'sales_compare': cat_sum['sales_compare'],
                'sales_yoy': calc_yoy(cat_sum['sales'], cat_sum['sales_compare']),
                'traffic': cat_sum['traffic'],
                'traffic_share': safe_div(cat_sum['traffic'], total_traffic) if total_traffic else None,
                'traffic_compare': cat_sum['traffic_compare'],
                'traffic_yoy': calc_yoy(cat_sum['traffic'], cat_sum['traffic_compare']),
                'is_subtotal': True,
            })

    # 精品总
    result.append({
        'category': '精品总',
        'tier': '',
        'sales': total_sales,
        'sales_share': 1.0,
        'sales_compare': total_sales_compare,
        'sales_yoy': calc_yoy(total_sales, total_sales_compare),
        'traffic': total_traffic,
        'traffic_share': 1.0,
        'traffic_compare': total_traffic_compare,
        'traffic_yoy': calc_yoy(total_traffic, total_traffic_compare),
        'is_subtotal': False,
        'is_grand_total': True,
    })

    return result


def build_brand_list(
    df: pd.DataFrame,
    mapping: Dict[str, Dict[str, str]],
) -> List[Dict[str, Any]]:
    """构建品牌视角明细数据"""
    brands = []
    for _, row in df.iterrows():
        sn = str(row['品牌SN'])
        tier_info = mapping.get(sn, {})
        tier = tier_info.get('tier', '')
        category = tier_info.get('category', '')
        # 无映射标签的品牌不展示
        if not tier:
            continue
        sales = parse_num(row['销售额(含拒退)']) 
        traffic = parse_num(row['累计曝光流量'])
        sales_compare = parse_num(row['销售额(含拒退)(对比)'])
        traffic_compare = parse_num(row['累计曝光流量(对比)'])
        # 无数据的不展示
        if (sales is None or sales == 0) and (traffic is None or traffic == 0):
            continue
        brands.append({
            'group': str(row['二级业绩部类']),
            'sn': sn,
            'brand': str(row['品牌名称'] or ''),
            'tier': tier,
            'category': category,
            'sales': sales or 0,
            'sales_compare': sales_compare or 0,
            'sales_yoy': calc_yoy(sales, sales_compare),
            'traffic': traffic or 0,
            'traffic_compare': traffic_compare or 0,
            'traffic_yoy': calc_yoy(traffic, traffic_compare),
        })
    return brands


def generate_audit(
    df_original: pd.DataFrame,
    df_clean: pd.DataFrame,
    mapping: Dict[str, Dict[str, str]],
    summary: List[Dict[str, Any]],
    brands: List[Dict[str, Any]],
    source_date: str,
) -> Dict[str, Any]:
    """生成审计报告"""
    total_sns = df_original['品牌SN'].nunique()
    clean_sns = df_clean['品牌SN'].nunique()
    mapped_sns = len(set(df_clean['品牌SN'].unique()) & set(mapping.keys()))
    sales_total = float(df_clean['销售额(含拒退)'].sum() or 0)
    traffic_total = float(df_clean['累计曝光流量'].sum() or 0)

    # 汇总核验
    summary_checks = {}
    if summary:
        total = summary[-1]
        summary_checks['精品总业绩'] = total['sales']
        summary_checks['精品总曝光'] = total['traffic']

        # 标品合计 + 类穿戴合计
        cat_sums = {}
        for s in summary:
            if s.get('is_subtotal'):
                cat_sums[s['category']] = s['sales']
        check_sum = sum(cat_sums.values())
        # 允许舍入误差
        summary_checks['标品+类穿戴'] = check_sum
        summary_checks['=精品总?'] = abs(check_sum - total['sales']) / max(total['sales'], 1) < 1e-6

        # 占比和
        share_sum = sum(s.get('sales_share') or 0 for s in summary if not s.get('is_subtotal') and not s.get('is_grand_total'))
        summary_checks['占比之和'] = share_sum
        summary_checks['占比=100%?'] = abs(share_sum - 1.0) < 1e-4

    return {
        'source_date': source_date,
        'generated_at': dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'source_file': str(df_original.name) if hasattr(df_original, 'name') else '',
        'data_quality': {
            'original_rows': len(df_original),
            'clean_rows': len(df_clean),
            'original_unique_sns': total_sns,
            'clean_unique_sns': clean_sns,
            'mapped_sns': mapped_sns,
            'mapping_coverage': safe_div(mapped_sns, clean_sns),
            'unmatched_with_data': clean_sns - mapped_sns,
            'excluded_historical_groups': len(HISTORICAL_GROUPS),
            'excluded_historical_rows': len(df_original) - len(df_clean),
        },
        'summary_verification': summary_checks,
        'brands_count': len(brands),
        'summary_rows': len(summary),
    }


def main():
    p = argparse.ArgumentParser(description='品牌分层MTD数据生成')
    p.add_argument('--vma', required=True, help='VMA下载的Excel文件路径')
    p.add_argument('--mapping', default=DEFAULT_MAPPING_SOURCE, help='货价监控表路径（含品牌分层映射O-T列）')
    p.add_argument('--output', default=OUTPUT_FILE, help='输出JSON路径')
    p.add_argument('--audit-dir', default='.', help='审计报告输出目录')
    p.add_argument('--source-date', default=None, help='数据日期，默认从文件名推断')
    args = p.parse_args()

    print(f'{"="*60}')
    print(f'品牌分层MTD 数据生成')
    print(f'VMA: {args.vma}')
    print(f'映射: {args.mapping}')
    print(f'输出: {args.output}')
    print(f'{"="*60}')

    # 1. 读取数据
    print('\n[1/4] 读取VMA数据...')
    df = read_vma_data(args.vma)
    df_original = df.copy()

    # 2. 读取映射
    print('\n[2/4] 读取品牌分层映射...')
    mapping = read_tier_mapping(args.mapping)

    # 3. 计算汇总
    print('\n[3/4] 计算汇总...')
    summary = calc_summary(df, mapping, matched_only=True)
    print(f'  汇总表: {len(summary)} 行')
    for s in summary:
        tier_label = s['tier'] or s['category']
        print(f'    {s["category"]:6s} {tier_label:6s} 业绩={s["sales"]:>12.1f} 曝光={s["traffic"]:>10.0f}')

    # 4. 构建品牌明细
    print('\n[4/4] 构建品牌明细...')
    brands = build_brand_list(df, mapping)
    print(f'  品牌明细: {len(brands)} 条记录')

    # 统计分标签
    tier_in_brands = {}
    for b in brands:
        t = b['tier'] or '(无标签)'
        tier_in_brands[t] = tier_in_brands.get(t, 0) + 1
    print(f'  品牌分层分布: {dict(sorted(tier_in_brands.items()))}')

    # 数据日期
    source_date = args.source_date
    if not source_date:
        # 从文件名或今天取
        from pathlib import Path
        import re
        # 文件名: vma_品牌分层数据8.11_xxx -> 2026-08-11
        m = re.search(r'数据(\d{1,2})\.(\d{1,2})', Path(args.vma).name)
        if m:
            source_date = f'2026-{int(m.group(1)):02d}-{int(m.group(2)):02d}'
        else:
            source_date = dt.datetime.now().strftime('%Y-%m-%d')

    # 输出
    output = {
        'schema_version': '2.0',
        'source_date': source_date,
        'updated_at': dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'source_file': str(Path(args.vma).name),
        'mapping_source': str(Path(args.mapping).name),
        'summary': summary,
        'brands': brands,
    }

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f'\n✅ 已生成: {output_path}')

    # 审计报告
    audit = generate_audit(df_original, df, mapping, summary, brands, source_date)
    audit['output_file'] = str(output_path)

    audit_path = Path(args.audit_dir) / f'品牌分层MTD_复核报告_{source_date.replace("-","")}.json'
    with open(audit_path, 'w', encoding='utf-8') as f:
        json.dump(audit, f, ensure_ascii=False, indent=2)
    print(f'✅ 审计报告: {audit_path}')

    # 校验
    print(f'\n{"="*60}')
    print('门禁校验')
    print(f'{"="*60}')
    print(f'  VMA原始行数: {len(df_original)}')
    print(f'  VMA清洗行数: {len(df)}')
    print(f'  映射覆盖率: {audit["data_quality"]["mapping_coverage"]*100:.1f}%')

    # 标品+类穿戴=精品总
    check = audit['summary_verification'].get('=精品总?', False)
    print(f'  标品合计+类穿戴合计=精品总: {"✅" if check else "❌"}')

    share_check = audit['summary_verification'].get('占比=100%?', False)
    print(f'  占比之和=100%: {"✅" if share_check else "❌"}')

    # 门禁
    if not check:
        print('\n❌ 门禁失败: 标品合计+类穿戴合计 ≠ 精品总')
        sys.exit(1)
    if not share_check:
        print('\n❌ 门禁失败: 占比之和 ≠ 100%')
        sys.exit(1)

    print('\n✅ 全部门禁通过，数据就绪')
    return 0


if __name__ == '__main__':
    sys.exit(main())