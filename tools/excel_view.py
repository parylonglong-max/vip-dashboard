#!/usr/bin/env python3
"""Excel-like mobile view builder for 精品货价监控数据看板.

Goal: preserve Sheet1's business layout as much as possible for mobile viewing.
Numbers are rendered in ten-thousand units (万) when values look like amounts/counts,
while rates/scores/pp/index fields preserve their natural scale.
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
import ast
import operator
import re

from openpyxl.utils.cell import column_index_from_string, range_boundaries

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 行号动态定位
import os as _os, sys as _sys
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..', '..', '..', 'vip-price-monitor-email-pipeline-self-contained', 'scripts', 'modules'))
try:
    from row_locator import find_module_range as _find_module_range, find_section_header_row as _find_header, find_subsection_header as _find_sub, find_next_module_row as _find_next
except ImportError:
    _find_module_range = None
    _find_header = None
    _find_sub = None
    _find_next = None

# 模块定义：(关键词, 是否有子区域)
_MODULE_DEFS = [
    ("自营销售", True),
    ("毛利", False),
    ("外网价指", True),
    ("内网折扣", True),
    ("六高", True),
    ("优质款", True),
    ("机采", True),
    ("五星价格力", True),
]

def _compute_section_ranges(ws):
    """动态计算各模块的行范围，返回 {模块关键词: (start, end)}"""
    ranges = {}
    if _find_module_range is None:
        return ranges
    for i, (kw, has_sub) in enumerate(_MODULE_DEFS):
        # 找模块标题行
        header_row = _find_header(ws, [kw], search_col=2, start_row=1, end_row=200)
        if header_row is None:
            continue
        # 找下一个模块的标题行作为结束
        next_kws = [_MODULE_DEFS[j][0] for j in range(i+1, len(_MODULE_DEFS))]
        end_row = 200
        for r in range(header_row+1, 201):
            val = str(ws.cell(r, 2).value or '').strip()
            for nk in next_kws:
                if nk in val:
                    end_row = r - 1
                    break
            else:
                continue
            break
        ranges[kw] = (header_row, end_row)
    return ranges

def _get_section_range(ws, module_kw, sub_section=None, fallback=None):
    """获取模块内某个子区域的行范围"""
    if _find_module_range is None:
        return fallback
    ranges = _compute_section_ranges(ws)
    mod_range = ranges.get(module_kw)
    if mod_range is None:
        return fallback
    start, end = mod_range
    if sub_section:
        sub_row = _find_sub(ws, [sub_section], search_col=2, start_row=start+1, end_row=end)
        if sub_row:
            # 子区域从标题行下一行开始到下一个子标题或模块结束
            next_sub = None
            for r in range(sub_row+1, end+1):
                val = str(ws.cell(r, 2).value or '').strip()
                if val in ('YTD', 'MTD', '历史月份', '历史月份得分'):
                    next_sub = r
                    break
            sub_end = next_sub - 1 if next_sub else end
            return (sub_row, sub_end)
    return (start, end)

SECTION_SPECS = [
    {"id": "self_sales_mtd", "title": "自营销售 · MTD", "range": (3, 12, 2, 11), "stickyCols": 1},
    {"id": "self_sales_history", "title": "自营销售 · YTD / 历史月份", "range": (15, 23, 2, 37), "stickyCols": 1},
    {"id": "gross_profit", "title": "毛利 · 单独更新", "range": (26, 38, 2, 11), "stickyCols": 1},
    {"id": "price_index_mtd", "title": "外网价指 · MTD", "range": (41, 51, 2, 15), "stickyCols": 1},
    {"id": "price_index_history", "title": "外网价指 · YTD / 历史月份得分", "range": (53, 62, 2, 79), "stickyCols": 1},
    {"id": "internal_discount", "title": "内网折扣 · MTD / YTD / 历史月份", "range": (65, 74, 2, 26), "stickyCols": 1},
    {"id": "six_high", "title": "六高 · MTD", "range": (77, 86, 2, 16), "stickyCols": 1},
    {"id": "quality_product_mtd", "title": "优质款 · MTD", "range": (89, 98, 2, 7), "stickyCols": 1},
    {"id": "quality_product_history", "title": "优质款 · YTD / 历史月份", "range": (100, 109, 2, 37), "stickyCols": 1},
    {"id": "machine_purchase_mtd", "title": "机采 · MTD", "range": (112, 118, 2, 9), "stickyCols": 1},
    {"id": "machine_purchase_history", "title": "机采 · YTD / 历史月份", "range": (121, 126, 2, 37), "stickyCols": 1},
    {"id": "price_power_mtd", "title": "五星价格力 & 大爆款效率 · MTD", "range": (129, 134, 2, 10), "stickyCols": 1},
    {"id": "price_power_history", "title": "五星价格力 & 大爆款效率 · YTD / 历史月份", "range": (137, 152, 2, 11), "stickyCols": 2},
]

# 动态范围映射：section_id -> (模块关键词, 子区域, fallback_range)
_DYNAMIC_RANGE_MAP = {
    "self_sales_mtd": ("自营销售", "MTD", (3, 12, 2, 11)),
    "self_sales_history": ("自营销售", "YTD", (15, 23, 2, 37)),
    "gross_profit": ("毛利", None, (26, 38, 2, 11)),
    "price_index_mtd": ("外网价指", "MTD", (41, 51, 2, 15)),
    "price_index_history": ("外网价指", "YTD", (53, 62, 2, 79)),
    "internal_discount": ("内网折扣", None, (65, 74, 2, 26)),
    "six_high": ("六高", "MTD", (77, 86, 2, 16)),
    "quality_product_mtd": ("优质款", "MTD", (89, 98, 2, 7)),
    "quality_product_history": ("优质款", "YTD", (100, 109, 2, 37)),
    "machine_purchase_mtd": ("机采", "MTD", (112, 118, 2, 9)),
    "machine_purchase_history": ("机采", "YTD", (121, 126, 2, 37)),
    "price_power_mtd": ("五星价格力", None, (129, 134, 2, 10)),
    "price_power_history": ("五星价格力", "YTD", (137, 152, 2, 11)),
}

def _resolve_section_ranges(ws):
    """动态计算所有section的行范围，替换硬编码。"""
    if _find_module_range is None:
        return  # 回退到硬编码
    for spec in SECTION_SPECS:
        sid = spec["id"]
        if sid not in _DYNAMIC_RANGE_MAP:
            continue
        mod_kw, sub, fallback = _DYNAMIC_RANGE_MAP[sid]
        dynamic = _get_section_range(ws, mod_kw, sub, fallback)
        if dynamic:
            r1, r2 = dynamic
            c1, c2 = fallback[2], fallback[3]  # 列范围保持不变
            spec["range"] = (r1, r2, c1, c2)

RATE_KEYWORDS = ("率", "同比", "完成率", "权重", "占比", "折扣", "价指", "价格指数", "指数")
SCORE_KEYWORDS = ("得分", "分")
DIFF_KEYWORDS = ("差值", "差距", "pp", "PP", "降幅", "vs", "VS")
AMOUNT_KEYWORDS = ("销售", "目标", "完成", "同期", "金额", "曝光", "商品数", "数量", "款数", "引进", "未引入", "总计", "APP销售", "外网加总")


def excel_date_to_str(val: Any) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(val))).strftime("%Y-%m-%d")
        except Exception:
            return str(val)
    return str(val)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"(NULL)", "NULL", "None", "nan"}:
        return ""
    return text.replace("\n", " ")


def merged_parent(ws, row: int, col: int) -> tuple[int, int] | None:
    """Return the top-left coordinate of the merged range containing row/col."""
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.min_row, merged.min_col
    return None


def cell_or_merged_value(ws, row: int, col: int):
    raw = ws.cell(row, col).value
    if raw is not None and normalize_text(raw) != "":
        return raw
    parent = merged_parent(ws, row, col)
    if parent and parent != (row, col):
        return ws.cell(parent[0], parent[1]).value
    return raw


def merged_span(ws, row: int, col: int) -> dict[str, Any] | None:
    """Return merge rendering metadata for frontend table rendering."""
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            if row == merged.min_row and col == merged.min_col:
                return {
                    "rowspan": merged.max_row - merged.min_row + 1,
                    "colspan": merged.max_col - merged.min_col + 1,
                    "covered": False,
                    "range": str(merged),
                }
            return {"covered": True, "range": str(merged)}
    return None


def get_display_raw(value_ws, formula_ws, row: int, col: int):
    """Return cached value first; fill merged cells; then fall back to formula/static workbook value."""
    raw = cell_or_merged_value(value_ws, row, col)
    if raw is not None and normalize_text(raw) != "":
        return raw
    fallback = cell_or_merged_value(formula_ws, row, col)
    if isinstance(fallback, str) and fallback.startswith("="):
        return None
    return fallback


def infer_col_context(value_ws, formula_ws, row: int, col: int, section_start_row: int) -> str:
    parts = []
    # Pick labels above/current column from the first 3 rows of the section.
    for r in range(section_start_row, min(section_start_row + 3, row + 1)):
        val = normalize_text(get_display_raw(value_ws, formula_ws, r, col))
        if val and val not in parts:
            parts.append(val)
    # include nearest left row label for row-based metrics
    row_label = normalize_text(get_display_raw(value_ws, formula_ws, row, 2))
    if row_label and row_label not in parts:
        parts.append(row_label)
    return " ".join(parts)


def should_show_as_percent(context: str, value: float) -> bool:
    # Excel stores most ratios as 0.x. Keep them as percentages when context says ratio/rate/weight/share.
    return any(k in context for k in ("率", "权重", "占比", "完成率", "同比")) and -5 <= value <= 5


def should_show_as_wan(context: str, value: float) -> bool:
    if any(k in context for k in RATE_KEYWORDS + SCORE_KEYWORDS + DIFF_KEYWORDS):
        return False
    if any(k in context for k in AMOUNT_KEYWORDS):
        return abs(value) >= 10000
    # Large raw values are likely amount/count columns; convert for mobile readability.
    return abs(value) >= 100000


def fmt_number(value: float, context: str, number_format: str = "") -> tuple[str, str]:
    fmt = (number_format or "").lower()
    if "%" in fmt:
        decimals = 0
        if "." in fmt:
            decimals = len(fmt.split(".", 1)[1].split("%", 1)[0])
        return f"{value * 100:.{decimals}f}%", "%"
    if should_show_as_percent(context, value):
        return f"{value * 100:.1f}%", "%"
    if should_show_as_wan(context, value):
        if abs(value) >= 100000000:
            yi = value / 100000000
            return f"{yi:,.2f}".rstrip("0").rstrip("."), "亿"
        wan = value / 10000
        if abs(wan) >= 100:
            return f"{wan:,.0f}", "万"
        return f"{wan:,.1f}", "万"
    if abs(value) >= 1000 and float(value).is_integer():
        return f"{value:,.0f}", ""
    if abs(value) >= 100:
        return f"{value:,.1f}".rstrip("0").rstrip("."), ""
    return f"{value:.1f}".rstrip("0").rstrip("."), ""



_BIN_OPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
}
_UNARY_OPS = {ast.UAdd: operator.pos, ast.USub: operator.neg}
_CELL_REF_RE = re.compile(r"(?<![A-Za-z0-9_])\$?([A-Z]{1,3})\$?(\d+)(?![A-Za-z0-9_])")
_RANGE_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+")


def _split_top_level_args(text: str) -> list[str]:
    args, current, depth, in_quote = [], [], 0, False
    quote_char = ""
    for ch in text:
        if ch in "'\"":
            if not in_quote:
                in_quote = True
                quote_char = ch
            elif quote_char == ch:
                in_quote = False
            current.append(ch)
            continue
        if not in_quote:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif ch == "," and depth == 0:
                args.append("".join(current).strip())
                current = []
                continue
        current.append(ch)
    args.append("".join(current).strip())
    return args


def _safe_eval_arithmetic(expr: str) -> float:
    tree = ast.parse(expr, mode="eval")

    def ev(node):
        if isinstance(node, ast.Expression):
            return ev(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return node.value
        if isinstance(node, ast.BinOp) and type(node.op) in _BIN_OPS:
            left = ev(node.left)
            right = ev(node.right)
            return _BIN_OPS[type(node.op)](left, right)
        if isinstance(node, ast.UnaryOp) and type(node.op) in _UNARY_OPS:
            return _UNARY_OPS[type(node.op)](ev(node.operand))
        raise ValueError(f"unsupported formula expression: {ast.dump(node)}")

    return float(ev(tree))


def _range_sum(value_ws, formula_ws, range_text: str, visited: set[tuple[int, int]]) -> float:
    min_col, min_row, max_col, max_row = range_boundaries(range_text.replace("$", ""))
    total = 0.0
    for rr in range(min_row, max_row + 1):
        for cc in range(min_col, max_col + 1):
            val = evaluate_cell_value(value_ws, formula_ws, rr, cc, visited)
            if isinstance(val, (int, float)):
                total += float(val)
    return total


def evaluate_formula(value_ws, formula_ws, formula: str, visited: set[tuple[int, int]]):
    expr = formula.strip()
    if expr.startswith("="):
        expr = expr[1:].strip()
    upper = expr.upper()

    if upper.startswith("IFERROR(") and expr.endswith(")"):
        inner = expr[8:-1]
        args = _split_top_level_args(inner)
        fallback = None
        if len(args) > 1:
            fallback = args[1].strip().strip('"').strip("'")
        try:
            return evaluate_formula(value_ws, formula_ws, "=" + args[0], visited)
        except Exception:
            return fallback if fallback is not None else ""

    if upper.startswith("IF(") and expr.endswith(")"):
        args = _split_top_level_args(expr[3:-1])
        if len(args) >= 3:
            cond = args[0]
            # Convert Excel comparisons into a Python boolean expression after resolving refs.
            def cond_ref(match):
                col_letters, row_s = match.group(1), match.group(2)
                rr = int(row_s)
                cc = column_index_from_string(col_letters)
                val = evaluate_cell_value(value_ws, formula_ws, rr, cc, visited)
                if isinstance(val, (int, float)):
                    return str(float(val))
                try:
                    return str(float(str(val).replace(",", "")))
                except Exception:
                    return "0"
            cond_expr = _CELL_REF_RE.sub(cond_ref, cond).replace("%", "/100").replace("=", "==")
            cond_expr = cond_expr.replace(">==", ">=").replace("<==", "<=").replace("!==", "!=")
            try:
                branch = args[1] if bool(_safe_eval_arithmetic(cond_expr)) else args[2]
            except Exception:
                branch = args[2]
            return evaluate_formula(value_ws, formula_ws, "=" + branch, visited)

    # Evaluate aggregate functions. Supports SUM(A1:B2), SUM(A1,B1), AVERAGE(...)
    def aggregate_repl(match):
        func = match.group(1).upper()
        inner = match.group(2)
        values = []
        for arg in _split_top_level_args(inner):
            arg = arg.strip().replace("$", "")
            if not arg:
                continue
            if ":" in arg and _RANGE_RE.fullmatch(arg):
                min_col, min_row, max_col, max_row = range_boundaries(arg)
                for rr in range(min_row, max_row + 1):
                    for cc in range(min_col, max_col + 1):
                        val = evaluate_cell_value(value_ws, formula_ws, rr, cc, visited)
                        if isinstance(val, (int, float)):
                            values.append(float(val))
            else:
                try:
                    values.append(float(evaluate_formula(value_ws, formula_ws, "=" + arg, visited)))
                except Exception:
                    val = None
                    m = _CELL_REF_RE.fullmatch(arg)
                    if m:
                        val = evaluate_cell_value(value_ws, formula_ws, int(m.group(2)), column_index_from_string(m.group(1)), visited)
                    if isinstance(val, (int, float)):
                        values.append(float(val))
        if not values:
            return "0"
        if func == "SUM":
            return str(sum(values))
        if func == "AVERAGE":
            return str(sum(values) / len(values))
        return "0"

    # Resolve innermost aggregate calls iteratively.
    for _ in range(10):
        new_expr = re.sub(r"(SUM|AVERAGE)\(([^()]+)\)", aggregate_repl, expr, flags=re.IGNORECASE)
        if new_expr == expr:
            break
        expr = new_expr

    # Replace cell refs with numeric values.
    def cell_repl(match):
        col_letters, row_s = match.group(1), match.group(2)
        rr = int(row_s)
        cc = column_index_from_string(col_letters)
        val = evaluate_cell_value(value_ws, formula_ws, rr, cc, visited)
        if isinstance(val, (int, float)):
            return str(float(val))
        if normalize_text(val) in {"", "-"}:
            return "0"
        try:
            return str(float(str(val).replace(",", "")))
        except Exception:
            return "0"

    expr = _CELL_REF_RE.sub(cell_repl, expr)
    # Remove simple Excel absolute markers and spaces.
    expr = expr.replace("$", "").replace(" ", "").replace("%", "/100")
    return _safe_eval_arithmetic(expr)


def evaluate_cell_value(value_ws, formula_ws, row: int, col: int, visited: set[tuple[int, int]] | None = None):
    # Fill merged cells from their top-left value. This is critical for mobile table views.
    parent = merged_parent(formula_ws, row, col) or merged_parent(value_ws, row, col)
    if parent and parent != (row, col):
        row, col = parent

    cached = value_ws.cell(row, col).value
    if cached is not None and normalize_text(cached) != "":
        return cached
    formula_or_value = formula_ws.cell(row, col).value
    if not (isinstance(formula_or_value, str) and formula_or_value.startswith("=")):
        return formula_or_value
    visited = visited or set()
    key = (row, col)
    if key in visited:
        return None
    visited.add(key)
    try:
        return evaluate_formula(value_ws, formula_ws, formula_or_value, visited)
    except Exception:
        return None
    finally:
        visited.discard(key)

def cell_to_view(value_ws, formula_ws, row: int, col: int, section_start_row: int) -> dict[str, Any]:
    raw = evaluate_cell_value(value_ws, formula_ws, row, col)
    coord = f"{get_column_letter(col)}{row}"
    formula = formula_ws.cell(row, col).value
    formula_text = formula if isinstance(formula, str) and formula.startswith("=") else ""
    merge = merged_span(formula_ws, row, col) or merged_span(value_ws, row, col)
    base: dict[str, Any] = {"coord": coord, "formula": formula_text}
    if merge:
        base["merge"] = merge
    if raw is None or normalize_text(raw) == "":
        base.update({"raw": None, "text": "", "unit": "", "type": "blank"})
        return base
    if isinstance(raw, (int, float)):
        cell = value_ws.cell(row, col)
        context = infer_col_context(value_ws, formula_ws, row, col, section_start_row)
        number_format = formula_ws.cell(row, col).number_format or cell.number_format or ""
        text, unit = fmt_number(float(raw), context, number_format)
        base.update({"raw": raw, "text": text, "unit": unit, "type": "number", "context": context, "numberFormat": number_format})
        return base
    text = normalize_text(raw)
    base.update({"raw": text, "text": text, "unit": "", "type": "text"})
    return base


def build_section(value_ws, formula_ws, spec: dict[str, Any]) -> dict[str, Any]:
    r1, r2, c1, c2 = spec["range"]
    rows = []
    for r in range(r1, r2 + 1):
        values = [cell_to_view(value_ws, formula_ws, r, c, r1) for c in range(c1, c2 + 1)]
        # Keep rows that contain at least one visible value.
        if any(cell["text"] for cell in values):
            rows.append({"excelRow": r, "cells": values})
    return {
        "id": spec["id"],
        "title": spec["title"],
        "range": f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}",
        "stickyCols": spec.get("stickyCols", 1),
        "rows": rows,
    }


def build_excel_view(excel_path: str | Path) -> dict[str, Any]:
    excel_path = Path(excel_path)
    value_wb = load_workbook(excel_path, data_only=True)
    formula_wb = load_workbook(excel_path, data_only=False)
    value_ws = value_wb["Sheet1"]
    formula_ws = formula_wb["Sheet1"]
    # 动态计算各模块行范围（替代硬编码）
    _resolve_section_ranges(value_ws)
    data_date = excel_date_to_str(evaluate_cell_value(value_ws, formula_ws, 1, 3))
    sections = [build_section(value_ws, formula_ws, spec) for spec in SECTION_SPECS]
    value_wb.close()
    formula_wb.close()
    return {
        "meta": {
            "title": "精品货价监控数据看板",
            "subtitle": "按 Excel Sheet1 原样呈现，数字统一优化为万单位",
            "sourceFile": excel_path.name,
            "sheet": "Sheet1",
            "dataDate": data_date,
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updatedAtHour": datetime.now().strftime("%Y-%m-%d %H点"),
            "versionNote": f"更新于 {datetime.now().strftime('%Y-%m-%d %H点')}",
        },
        "sections": sections,
    }


if __name__ == "__main__":
    import argparse, json
    parser = argparse.ArgumentParser()
    parser.add_argument("excel_path")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    data = build_excel_view(args.excel_path)
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(out)
